# This Python file uses the following encoding: utf-8
from os import name
import pdfplumber
import pandas as pd
import openpyxl
import openpyxl.styles
from openpyxl.styles import Font, Border
import time
from datetime import date
import json

json_file_name ='templates/config.json'
def pdf_to_txt(file):
    """ Converting every line in the pdf into a line into an excel it created with the name of the pdf + "result" """
    
    json_file = open(json_file_name, encoding="utf8")
    json_data = json.load(json_file)
    excel_file = pd.DataFrame()
    excel_file_name = file[:-4] +".xlsx"
    excel_file.to_excel(excel_file_name)
    book = openpyxl.load_workbook(json_data['path'] + excel_file_name)
    sheet = book.active

    sheet.sheet_view.rightToLeft = True    

    have_found_file_type = False

    people_row_count = json_data['excel_row_information_start']
    company_row_count = json_data['excel_row_information_start']
    passport_row_count = json_data['excel_row_information_start']

    file_type = 0
    added_row = 0


    pdf_file_name= file[:-24] +".pdf"
    with pdfplumber.open(pdf_file_name) as pdf:
        for page in pdf.pages:
            val= str(page)[1:]
            val=val[:-1]
            page_num= val[5:]
            page_amount= " out of  "+(str(pdf.pages)[len(str(pdf.pages))-6:])[:-2][-2:]
            val = val+ page_amount
            write_data_in_information_file(val,file[:-4]+".txt")
            for line in page.extract_text().split('\n'):
                if have_found_file_type:                   
                    added_row = line_information_extractor(
                        line, file_type, sheet, people_row_count, company_row_count, passport_row_count,page_num)
                    if added_row == 1:
                        people_row_count += 1
                    elif added_row == 2:
                        company_row_count += 1
                    elif added_row == 3:
                        passport_row_count += 1
                else:
                    if find_file_type(line, sheet) == 1:
                        have_found_file_type = True
                        file_type = 1
                    elif find_file_type(line, sheet) == 2:
                        have_found_file_type = True
                        file_type = 2
                   

        # Adding titles
    write_excel_titles(sheet)
    
    # Saving the excel
    new_excel_file_name=excel_file_name[:-5] + " result.xlsx"
    book.title=new_excel_file_name
    book.save(new_excel_file_name)   

    
    # Adding the information to the information file
    write_data_in_information_file("Finished extracting "+ file[:-4],file[:-4]+".txt")
    


def line_information_extractor(info, type_of_file, sheet, people_row_count, company_row_count, passport_row_count,page):
    """getting a line and checking if a certain information is in it then writing it in the excel"""
    json_file = open(json_file_name, encoding="utf8")
    json_data = json.load(json_file)

    row_added = 0

    if isinstance(info, str):
        # Checking if there`s ID in the line
        if json_data['hebrew_ID'] in info:
            info = info + " "
            info = " ".join(info.split())
            info = " " + info
            print(info)

            if type_of_file == 1:
                id_value= get_ID_from_sentence(info)
               
                id_value = id_value.replace(" ", "")

                sheet.cell(row=people_row_count, column=1).value = id_value
                sheet.cell(row=people_row_count, column=1).font = Font(size=11, bold=False)
    


                # Find the name and putting it in the excel by certain distance from the ID and the reason
                # name_value = info[info.find(json_data['hebrew_ID']) + 3:info.find(
                #     json_data['hebrew_ID']) + find_name_shared_homes(info)][::-1]
                name_value = get_ID_name_from_sentence(info)[::-1]                   
                
                sheet.cell(row=people_row_count, column=2).value = name_value
                sheet.cell(row=people_row_count, column=2).font = Font(size=11, bold=False)

                # Printing for debugging
                print(id_value)
                print(name_value)

            if type_of_file == 2:
                # Find the ID and putting it in the excel (ID is always in the start of the file, very simple check)
                #id_value = info[:info.find(json_data['hebrew_ID'])]     
                id_value= id_value= get_ID_from_sentence(info)
                        
                id_value = id_value.replace(" ", "")
                
                sheet.cell(row=people_row_count, column=1,).value = id_value
                sheet.cell(row=people_row_count, column=1).font = Font(size=11, bold=False)

                # Find the name and putting it in the excel by certain distance from the ID and the reason
                # name_value = info[info.find(json_data['hebrew_ID']) + 3:info.find(
                #     json_data['hebrew_ID']) + find_name_shared_rights(info)][::-1]   
                #       
                name_value = get_ID_name_from_sentence(info)                    

                sheet.cell(row=people_row_count, column=2).value = name_value
                sheet.cell(row=people_row_count, column=2).font = Font(size=11, bold=False)

                # Printing for debugging
                print(id_value)
                print(name_value)

            # Adding 1 to the index of where the program will write
            sheet.cell(row=people_row_count, column=3).value = page
            sheet.cell(row=people_row_count, column=3).font = Font(size=11, bold=False)
            row_added = 1
        # Checking if there`s company and not mortgage in the line
        elif json_data['hebrew_Company'] in info and json_data['hebrew_Mortgage'] not in info:
            info += " "
            info = " " + info
            info = " ".join(info.split())
            print(info)

            if type_of_file == 1:
                # Find the company ID and putting it in the excel               
                company_id_value= get_ID_from_sentence(info)               
                company_id_value = company_id_value.replace(" ", "")

                sheet.cell(row=company_row_count, column=5).value = company_id_value
                sheet.cell(row=company_row_count, column=5).font = Font(size=11, bold=False)

                # Find the name and putting it in the excel by certain distance from the ID and the reason
                company_name_value = info[info.find(json_data['hebrew_Company']) + 4:info.find(
                    json_data['hebrew_Company']) + find_company_name_shared_homes(info)][::-1]
                sheet.cell(row=company_row_count, column=6).value = company_name_value
                sheet.cell(row=company_row_count, column=6).font = Font(size=11, bold=False)

                # Printing for debugging
                print(company_name_value)
                print(company_id_value)

            if type_of_file == 2:
                # Find the company ID and putting it in the excel             
                company_id_value= get_ID_from_sentence(info)               
                company_id_value = company_id_value.replace(" ", "")
                
                sheet.cell(row=company_row_count, column=5).value = company_id_value
                sheet.cell(row=company_row_count, column=5).font = Font(size=11, bold=False)

                # Find the name and putting it in the excel by certain distance from the ID and the reason
                company_name_value = info[info.find(json_data['hebrew_Company']) + 4:info.find(
                    json_data['hebrew_Company']) + find_company_name_shared_rights(info)][::-1]
                sheet.cell(row=company_row_count, column=6).value = company_name_value
                sheet.cell(row=company_row_count, column=6).font = Font(size=11, bold=False)

                # Printing for debugging
                print(company_name_value)
                print(company_id_value)

            sheet.cell(row=company_row_count, column=7).value = page
            sheet.cell(row=company_row_count, column=7).font = Font(size=11, bold=False)

            # Adding 1 to the index of where the program will write
            row_added = 2
        # Checking if there`s passport and not mortgage in the line
        elif json_data['hebrew_passport'] in info:
            info += " "
            info = " " + info
            info = " ".join(info.split())
            print(info)

            if type_of_file == 1:
                # Find the passport and putting it in the excel (more complicated check, ID comes in multiple lengths)
                for i in range(9, 14):
                    if " " in info[
                              info.find(json_data['hebrew_passport']) - i:info.find(json_data['hebrew_passport']) - 1]:
                        passport_value = info[info.find(json_data['hebrew_passport']) - (i-1):info.find(
                            json_data['hebrew_passport']) - 1]
                        break

                sheet.cell(row=passport_row_count, column=8).value = passport_value
                sheet.cell(row=passport_row_count, column=8).font = Font(size=11, bold=False)

                # Find the name and putting it in the excel by certain distance from the passport and the reason
                passport_name_value = info[info.find(json_data['hebrew_passport']) + 5:info.find(
                    json_data['hebrew_passport']) + find_passport_name_shared_homes(info)][::-1]
                sheet.cell(row=passport_row_count, column=9).value = passport_name_value
                sheet.cell(row=passport_row_count, column=9).font = Font(size=11, bold=False)

                # Printing for debugging
                print(passport_value)
                print(passport_name_value)
            # Adding 1 to the index of where the program will write
            sheet.cell(row=passport_row_count, column=10).value = page
            sheet.cell(row=passport_row_count, column=10).font = Font(size=11, bold=False)
            row_added = 3

    return row_added


def write_data_in_information_file(value,filename):
    """writes in the information file the file name, and time of executing"""
    print(filename)
    information_file = open(filename, "a")
    information_file.write(value + " " + str(date.today().strftime("%d/%m/%Y")) + " " + str(time.strftime("%H:%M:%S", time.localtime()))+"\n")
    information_file.close()
    write_data_in_information_file1(value)

def write_data_in_information_file1(value):    
    filename = "Information.txt"
    information_file = open(filename, "a")
    information_file.write(value + " " + str(date.today().strftime("%d/%m/%Y")) + " " + str(time.strftime("%H:%M:%S", time.localtime()))+"\n")
    information_file.close()


def write_excel_titles(sheet):
    """gets a sheet and writes titles in it"""
    sheet.cell(row=1, column=1).value = "ת.ז"
    sheet.cell(row=1, column=1).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=2).value = "שם"
    sheet.cell(row=1, column=2).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=3).value = "מספר עמוד"
    sheet.cell(row=1, column=3).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=5).value = "מספר"
    sheet.cell(row=1, column=5).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=6).value = "שם חברה"
    sheet.cell(row=1, column=6).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=7).value = "מספר עמוד"
    sheet.cell(row=1, column=7).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=8).value = "מספר דרכון"
    sheet.cell(row=1, column=8).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=9).value = "שם"
    sheet.cell(row=1, column=9).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=10).value = "מספר עמוד"
    sheet.cell(row=1, column=10).font = Font(size=11, bold=True)


def clear_excel_cell(cell):
    """receive excel cell and clears it"""
    cell.value = None
    cell.font = Font(size=11, bold=False)
    cell.border = Border()


def write_file_type_in_excel(file_type, sheet):
    """gets sheet and file-type and writes it in the sheet"""
    sheet.cell(row=1, column=11).value = "סוג קובץ:"
    sheet.cell(row=1, column=11).font = Font(size=11, bold=True)
    sheet.cell(row=2, column=11).value = file_type
    sheet.cell(row=2, column=11).font = Font(size=11, bold=False)


def find_name_shared_rights(info):
    """Returning the distance of the name from the ID"""
    json_file = open(json_file_name, encoding="utf8")
    json_data = json.load(json_file)

    length = 3

    info = info[info.find(json_data['hebrew_ID']) + 3:]

    length += info.find(" ") + 1
    info = info[info.find(" ") + 1:]

    info = info[::-1]
    print(info)
    info = " ".join(info.split())
    info = info[name_reason_filtering_shared_rights(info):]

    info = " ".join(info.split())
    print(info)
    length += len(info)

    return length


def name_reason_filtering_shared_rights(info):
    """Returning the string without the reason in order to find the name"""
    json_file = open(json_file_name, encoding="utf8")
    json_data = json.load(json_file)

    index_of_reason = 0

    for reason in json_data['possible_name_reasons']:
        if reason in info:
            index_of_reason = info.find(reason) + len(reason)
            break
    json_file.close()
    return index_of_reason


def name_reason_filtering_shared_homes(info):
    """Returning the string without the reason in order to find the name"""

    json_file = open(json_file_name, encoding="utf8")
    json_data = json.load(json_file)

    for reason in json_data['possible_name_reasons']:
        if reason in info:
            info = info.replace(reason, "")
            break
    json_file.close()
    return info


def find_name_shared_homes(info):
    """Returning the distance of the name from the ID"""
    json_file = open(json_file_name, encoding="utf8")
    json_data = json.load(json_file)

    length = 3
    info = info[info.find(json_data['hebrew_ID']) + 3:]

    length += info.find(" ") + 1
    info = info[info.find(" ") + 1:]

    info = info[::-1]
    print(info)

    info = name_reason_filtering_shared_homes(info)

    info = " ".join(info.split())
    print(info)
    length += len(info)

    return length


def find_passport_name_shared_homes(info):
    """Returning the distance of the name from the passport ID"""
    json_file = open(json_file_name, encoding="utf8")
    json_data = json.load(json_file)

    length = 5
    info = info[info.find(json_data['hebrew_passport']) + 5:]

    info = " ".join(info.split())

    length += info.find(" ") + 1
    info = info[info.find(" ") + 1:]

    info = " ".join(info.split())

    info = info[::-1]
    print(info)

    info = name_reason_filtering_shared_homes(info)

    info = " ".join(info.split())
    print(info)
    length += len(info)

    return length


def find_company_name_shared_homes(info):
    """Returning the distance of the name from the company ID"""
    json_file = open(json_file_name, encoding="utf8")
    json_data = json.load(json_file)

    length = 4
    info = info[info.find(json_data['hebrew_Company']) + 4:]

    length += info.find(" ") + 1
    info = info[info.find(" ") + 1:]

    info = info[::-1]
    print(info)

    for reason in json_data['possible_company_name_reasons']:
        if reason in info:
            info = info.replace(reason, "")
    json_file.close()
    info = " ".join(info.split())

    print(info)
    length += len(info)

    return length


def find_company_name_shared_rights(info):
    """Returning the distance of the name from the company ID"""
    json_file = open(json_file_name, encoding="utf8")
    json_data = json.load(json_file)

    length = 4
    info = info[info.find(json_data['hebrew_Company']) + 4:]

    length += info.find(" ") + 1
    info = info[info.find(" ") + 1:]

    info = info[::-1]
    print(info)

    for reason in json_data['possible_company_name_reasons']:
        if reason in info:
            info = info[:info.find(reason)]
    json_file.close()
    info = " ".join(info.split())

    print(info)
    length += len(info)

    return length


def find_file_type(info, sheet):
    """Returning the type of the file presented by numbers"""
    if "םיפתושמ םיתב" in info:
        print(info)
        write_file_type_in_excel("בתים משותפים", sheet)
        return 1
    elif "תויוכזה סקנפמ" in info:
        print(info)
        write_file_type_in_excel("פנקס זכויות", sheet)
        return 2
    else:
        return 0


def get_ID_from_sentence(sentence):
    words = sentence.split()
    for word in words:
        if(len(word)>5):
            if(word.isnumeric() or ("/" not in word and "-" in word)):
                return word


def get_ID_name_from_sentence(info):
    """Returning the distance of the name from the ID"""
    json_file = open(json_file_name, encoding="utf8")
    json_data = json.load(json_file)

    info = info.replace(json_data['hebrew_ID'],"")

    info=info[::-1]

    for reason in json_data['possible_name_reasons']:
        if reason in info:
            info=info.replace(reason,"")
    
    for word in info:
        if(word.isdigit() or '/' in word):
            info= info.replace(word, "")

    for count in range(0, len(info)):
        if(info[count]==')'):
            info = info[:count] + '(' + info[count+1:]
        elif(info[count]=='('):
            info = info[:count] + ')' + info[count+1:]

   
    info = info.replace("  ", "")
    for a in range(0,2):
        if(info[len(info)-1] ==' ' or info[len(info)-1] =='-'):
            info = info[:-1]            
        elif(info[0] == ' ' or info[0] == '-'):
            info = info[1:]          
       
    info=info[::-1]

    return info
   
#get_ID_name_from_sentence(" 17728/2013/1 1 / 2 300812641 ז.ת הנוי רואמ רכמ")
#pdf_to_txt('352.pdf')
#print(find_passport_name_shared_homes(info))
#print(info[info.find("ןוכרד") - 10:info.find("ןוכרד") - 1])
#print(info[info.find("ןוכרד") + 5:info.find("ןוכרד") + find_passport_name_shared_homes(info)][::-1])


