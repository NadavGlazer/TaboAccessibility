from os import kill
from openpyxl.cell import cell
import tabula
import openpyxl
import openpyxl.styles
from openpyxl.styles import Font, Border, Alignment
import pandas as pd
import json
import time
from datetime import date


def extract_data_from_pdf(filename, file_type):
    """Gets pdf file name and extracts the data from it""" 
    new_excel_file_name = filename[:-4] + "_result.xlsx"
    text_file_name = filename[:-4] + "_text.txt"
    write_data_in_file("Extracting " + filename[:-4], filename[:-4] + ".txt")

    tabula.convert_into(filename, text_file_name, pages = "all")
    with open(text_file_name , "r", encoding="utf-8") as text_file:
        lines = text_file.readlines()
    
    if file_type == "SharedHomes":
        is_shared_homes = True
    elif file_type == "RightsNotepad":    
        is_shared_rights = True
    else:
        is_bills_notepad = True


    specific_info = []
    for line in lines:    
        if not line.replace("''''","").replace('""', "").replace(",", "").isspace():
            current_line =line.replace("\n", "").split(",")    

            isNot = True
            counter = 0
            for item in current_line:
                if item == "דרגה:ראשונה" or item == "ראשונה"  or "שניה" in item or item== "שלישית" or item == "רביעית"  or "הערות" in item or "החלק" in item or "משכנתה" in item:
                    isNot = False
                if not item == '':
                    counter += 1
        
            if counter < 5 :
                isNot= False
            if len(current_line) >5:  
                if current_line[5] == '' and not current_line[4] == '':
                    current_line= current_line[:5] + current_line[6:]     

            if isNot:            
                specific_info.append(current_line)        
  
    excel_file = pd.DataFrame()
    excel_file.to_excel(new_excel_file_name)
    book = openpyxl.load_workbook(new_excel_file_name)

    sheet = book.active

    sheet.sheet_view.rightToLeft = True

    specific_info = specific_info[2:]
    counter = 2
    for item in specific_info:
        item = get_array_item_type(item)   
    
        for i in range(0, len(item)):
            print(item[i])
            sheet.cell(row=counter, column=i +1).value = item[i].strip()
        counter += 1
    
    write_excel_titles(sheet)
    write_file_type_in_excel(file_type, sheet)

    book.save(new_excel_file_name)
    write_data_in_file("Finished extracting " + filename[:-4], filename[:-4] + ".txt")

def get_array_item_type(line):
    """Gets specific line and returns if its an id, name...."""

    counter = 0
    for item in line:        
        if item == "ת.ז" or item == "חברה" or item == "דרכון":         
            break
            
        counter += 1
    if counter == 3 or counter == len(line):
        return line        
    else:
        return line[:1] + [""] + line[1:]

def write_data_in_file(value, filename):    
    """writes in the information file the file name, and time of executing"""
    print(filename)
    information_file = open(filename, "a")
    information_file.write(
        value
        + " "
        + str(date.today().strftime("%d/%m/%Y"))
        + " "
        + str(time.strftime("%H:%M:%S", time.localtime()))
        + "\n"
    )
    information_file.close()
    write_data_in_information_file(value)

def write_data_in_information_file(value):
    """gets information and writes it in the information file"""
    filename = "Information.txt"
    information_file = open(filename, "a")
    information_file.write(
        value
        + " "
        + str(date.today().strftime("%d/%m/%Y"))
        + " "
        + str(time.strftime("%H:%M:%S", time.localtime()))
        + "\n"
    )
    information_file.close()

def write_excel_titles(sheet):
    """gets a sheet and writes titles in it"""
    sheet.cell(row=1, column=1).value = "חלקה"
    sheet.cell(row=1, column=1).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=2).value = "חלק"
    sheet.cell(row=1, column=2).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=3).value = "מספר זיהוי"
    sheet.cell(row=1, column=3).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=4).value = "סוג זיהוי"
    sheet.cell(row=1, column=4).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=5).value = "שם"
    sheet.cell(row=1, column=5).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=6).value = "מהות הפעולה "
    sheet.cell(row=1, column=6).font = Font(size=11, bold=True)

def write_file_type_in_excel(file_type, sheet):
    """gets sheet and file-type and writes it in the sheet"""
    sheet.cell(row=1, column=11).value = "סוג קובץ:"
    sheet.cell(row=1, column=11).font = Font(size=11, bold=True)
    sheet.cell(row=2, column=11).value = file_type
    sheet.cell(row=2, column=11).font = Font(size=11, bold=False)

def filter_full_file_to_array_shared_homes(lines):
    needed_lines = []
    for line in lines:    
        if not line.replace("''''","").replace('""', "").replace(",", "").isspace():
            current_line =line.replace("\n", "").split(",")    

            isNot = True
            counter = 0
            for item in current_line:
                if item == "דרגה:ראשונה" or item == "ראשונה"  or "שניה" in item or item== "שלישית" or item == "רביעית"  or "הערות" in item or "החלק" in item or "משכנתה" in item:
                    isNot = False
                if not item == '':
                    counter += 1
        
            if counter < 5 :
                isNot= False
            if len(current_line) >5:  
                if current_line[5] == '' and not current_line[4] == '':
                    current_line= current_line[:5] + current_line[6:]     

            if isNot:            
                needed_lines.append(current_line)
    return needed_lines  