import pdfplumber
import pandas as pd
import openpyxl
import openpyxl.styles
from openpyxl.styles import Font, Border
import time
from datetime import date
import json


def pdf_to_txt(file):
    """ Converting every line in the pdf into a line into an excel it created with the name of the pdf + "result" """
    json_file = open('config.json', encoding="utf8")
    json_data = json.load(json_file)

    excel_file = pd.DataFrame()
    excel_file_name = file[:-4] + " result.xlsx"
    excel_file.to_excel(excel_file_name)
    book = openpyxl.load_workbook(json_data['path'] + excel_file_name)
    sheet = book.active

    have_found_file_type = False

    people_row_count = json_data['excel_row_information_start']
    company_row_count = json_data['excel_row_information_start']
    passport_row_count = json_data['excel_row_information_start']

    file_type = 0
    added_row = 0
    lines = []
    print(str(time.strftime("%H:%M:%S", time.localtime())))
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            for line in text.split('\n'):
                lines.append(line)
    print(str(time.strftime("%H:%M:%S", time.localtime())))

    for line in lines:
        if not have_found_file_type:
            if find_file_type(line, sheet) == 1:
                have_found_file_type = True
                file_type = 1
            elif find_file_type(line, sheet) == 2:
                have_found_file_type = True
                file_type = 2
        else:
            added_row = line_information_extractor(
                line, file_type, sheet, people_row_count, company_row_count, passport_row_count)
            if added_row == 1:
                people_row_count += 1
            elif added_row == 2:
                company_row_count += 1
            elif added_row == 3:
                passport_row_count += 1
    print(str(time.strftime("%H:%M:%S", time.localtime())))

    # Adding titles
    write_excel_titles(sheet)

    # Saving the excel
    book.save(excel_file_name)

    # Adding the information to the information file
    write_data_in_information_file(file[:-4])


def line_information_extractor(info, type_of_file, sheet, people_row_count, company_row_count, passport_row_count):
    """getting a line and checking if a certain information is in it then writing it in the excel"""
    json_file = open('config.json', encoding="utf8")
    json_data = json.load(json_file)

    row_added = 0

    if isinstance(info, str):

        # Checking if there`s ID in the line
        if json_data['hebrew_ID'] in info:
            info = info + " "
            info = " ".join(info.split())
            info = " " + info
            print(info)

            if type_of_file == 1:
                # Find the ID and putting it in the excel (more complicated check, ID comes in multiple lengths)
                for i in range(6, 14):
                    if " " in info[info.find(json_data['hebrew_ID']) - i:info.find(json_data['hebrew_ID']) - 1]:
                        id_value = info[info.find(json_data['hebrew_ID']) - (i-1):info.find(json_data['hebrew_ID']) - 1]
                        break
                sheet.cell(row=people_row_count, column=1).value = id_value
                sheet.cell(row=people_row_count, column=1).font = Font(size=11, bold=False)

                # Find the name and putting it in the excel by certain distance from the ID and the reason
                name_value = info[info.find(json_data['hebrew_ID']) + 3:info.find(
                    json_data['hebrew_ID']) + find_name_shared_homes(info)][::-1]
                sheet.cell(row=people_row_count, column=2).value = name_value
                sheet.cell(row=people_row_count, column=2).font = Font(size=11, bold=False)

                # Printing for debugging
                print(id_value)
                print(name_value)

            if type_of_file == 2:
                # Find the ID and putting it in the excel (ID is always in the start of the file, very simple check)
                id_value = info[:info.find(json_data['hebrew_ID'])]
                sheet.cell(row=people_row_count, column=1).value = id_value
                sheet.cell(row=people_row_count, column=1).font = Font(size=11, bold=False)

                # Find the name and putting it in the excel by certain distance from the ID and the reason
                name_value = info[info.find(json_data['hebrew_ID']) + 3:info.find(
                    json_data['hebrew_ID']) + find_name_shared_rights(info)][::-1]
                sheet.cell(row=people_row_count, column=2).value = name_value
                sheet.cell(row=people_row_count, column=2).font = Font(size=11, bold=False)

                # Printing for debugging
                print(id_value)
                print(name_value)

            # Adding 1 to the index of where the program will write
            row_added = 1
        # Checking if there`s company and not mortgage in the line
        elif json_data['hebrew_Company'] in info and json_data['hebrew_Mortgage'] not in info:
            info += " "
            info = " " + info
            info = " ".join(info.split())
            print(info)

            if type_of_file == 1:
                # Find the company ID and putting it in the excel
                company_id_value = info[info.find(json_data['hebrew_Company']) - 10:info.find(
                    json_data['hebrew_Company']) - 1]
                sheet.cell(row=company_row_count, column=4).value = company_id_value
                sheet.cell(row=company_row_count, column=4).font = Font(size=11, bold=False)

                # Find the name and putting it in the excel by certain distance from the ID and the reason
                company_name_value = info[info.find(json_data['hebrew_Company']) + 4:info.find(
                    json_data['hebrew_Company']) + find_company_name_shared_homes(info)][::-1]
                sheet.cell(row=company_row_count, column=5).value = company_name_value
                sheet.cell(row=company_row_count, column=5).font = Font(size=11, bold=False)

                # Printing for debugging
                print(company_name_value)
                print(company_id_value)

            if type_of_file == 2:
                # Find the company ID and putting it in the excel
                company_id_value = info[info.find(json_data['hebrew_Company']) - 10:info.find(
                    json_data['hebrew_Company']) - 1]
                sheet.cell(row=company_row_count, column=4).value = company_id_value
                sheet.cell(row=company_row_count, column=4).font = Font(size=11, bold=False)

                # Find the name and putting it in the excel by certain distance from the ID and the reason
                company_name_value = info[info.find(json_data['hebrew_Company']) + 4:info.find(
                    json_data['hebrew_Company']) + find_company_name_shared_rights(info)][::-1]
                sheet.cell(row=company_row_count, column=5).value = company_name_value
                sheet.cell(row=company_row_count, column=5).font = Font(size=11, bold=False)

                # Printing for debugging
                print(company_name_value)
                print(company_id_value)

            # Adding 1 to the index of where the program will write
            row_added = 2
        # Checking if there`s passport and not mortgage in the line
        elif json_data['hebrew_passport'] in info:
            info += " "
            info = " " + info
            info = " ".join(info.split())
            print(info)

            if type_of_file == 1:
                # Find the passport and putting it in the excel (more complicated check, ID comes in multiple lengths)
                for i in range(9, 14):
                    if " " in info[
                              info.find(json_data['hebrew_passport']) - i:info.find(json_data['hebrew_passport']) - 1]:
                        passport_value = info[info.find(json_data['hebrew_passport']) - (i-1):info.find(
                            json_data['hebrew_passport']) - 1]
                        break

                sheet.cell(row=passport_row_count, column=7).value = passport_value
                sheet.cell(row=passport_row_count, column=7).font = Font(size=11, bold=False)

                # Find the name and putting it in the excel by certain distance from the passport and the reason
                passport_name_value = info[info.find(json_data['hebrew_passport']) + 5:info.find(
                    json_data['hebrew_passport']) + find_passport_name_shared_homes(info)][::-1]
                sheet.cell(row=passport_row_count, column=8).value = passport_name_value
                sheet.cell(row=passport_row_count, column=8).font = Font(size=11, bold=False)

                # Printing for debugging
                print(passport_value)
                print(passport_name_value)
            # Adding 1 to the index of where the program will write
            row_added = 3

    return row_added


def file_information_extractor(excel_file, file_name):
    """Gets the excel file and checks for the information and writes the data in the excel"""
    json_file = open('config.json', encoding="utf8")
    json_data = json.load(json_file)

    book = openpyxl.load_workbook(json_data['path'] + excel_file)
    sheet = book.active

    people_row_count = json_data['excel_row_information_start']
    company_row_count = json_data['excel_row_information_start']
    passport_row_count = json_data['excel_row_information_start']

    #A and B row in the excel, have different data at first
    a_row_in_excel = sheet['A']
    b_row_in_excel = sheet['B']

    #Finds the type of file and writing it on the excel
    type_of_file = 0
    for cell in b_row_in_excel:
        if isinstance(cell.internal_value, str):
            type_of_file = find_file_type(cell.internal_value, sheet)
            if type_of_file == 1 or type_of_file == 2:
                print(type_of_file)
                #break if type of file were found
                break

    #Resetting the design on the cells
    for cell in a_row_in_excel:
        clear_excel_cell(cell)

    #Checking for the information
    for cell in b_row_in_excel:
        info = cell.internal_value
        clear_excel_cell(cell)

        if isinstance(info, str):

            #Checking if there`s ID in the line
            if json_data['hebrew_ID'] in info:
                info = info + " "
                info = " ".join(info.split())
                info = " " + info
                print(info)

                if type_of_file == 1:
                    # Find the ID and putting it in the excel (more complicated check, ID comes in multiple lengths)
                    if " " in info[info.find(json_data['hebrew_ID']) - 9:info.find(json_data['hebrew_ID']) - 1]:
                        id_value = info[info.find(json_data['hebrew_ID']) - 8:info.find(json_data['hebrew_ID']) - 1]
                    elif " " in info[info.find(json_data['hebrew_ID']) - 10:info.find(json_data['hebrew_ID']) - 1]:
                        id_value = info[info.find(json_data['hebrew_ID']) - 9:info.find(json_data['hebrew_ID']) - 1]
                    elif " " in info[info.find(json_data['hebrew_ID']) - 11:info.find(json_data['hebrew_ID']) - 1]:
                        id_value = info[info.find(json_data['hebrew_ID']) - 10:info.find(json_data['hebrew_ID']) - 1]
                    elif " " in info[info.find(json_data['hebrew_ID']) - 12:info.find(json_data['hebrew_ID']) - 1]:
                        id_value = info[info.find(json_data['hebrew_ID']) - 11:info.find(json_data['hebrew_ID']) - 1]
                    elif " " in info[info.find(json_data['hebrew_ID']) - 13:info.find(json_data['hebrew_ID']) - 1]:
                        id_value = info[info.find(json_data['hebrew_ID']) - 12:info.find(json_data['hebrew_ID']) - 1]
                    else:
                        id_value = info[info.find(json_data['hebrew_ID']) - 13:info.find(json_data['hebrew_ID']) - 1]
                    sheet.cell(row=people_row_count, column=1).value = id_value
                    sheet.cell(row=people_row_count, column=1).font = Font(size=11, bold=False)

                    # Find the name and putting it in the excel by certain distance from the ID and the reason
                    name_value = info[info.find(json_data['hebrew_ID']) + 3:info.find(json_data['hebrew_ID']) + find_name_shared_homes(info)][::-1]
                    sheet.cell(row=people_row_count, column=2).value = name_value
                    sheet.cell(row=people_row_count, column=2).font = Font(size=11, bold=False)

                    # Printing for debugging
                    print(id_value)
                    print(name_value)

                if type_of_file == 2:
                    #Find the ID and putting it in the excel (ID is always in the start of the file, very simple check)
                    id_value = info[:info.find(json_data['hebrew_ID'])]
                    sheet.cell(row=people_row_count, column=1).value = id_value
                    sheet.cell(row=people_row_count, column=1).font = Font(size=11, bold=False)

                    #Find the name and putting it in the excel by certain distance from the ID and the reason
                    name_value = info[info.find(json_data['hebrew_ID']) + 3:info.find(json_data['hebrew_ID']) + find_name_shared_rights(info)][::-1]
                    sheet.cell(row=people_row_count, column=2).value = name_value
                    sheet.cell(row=people_row_count, column=2).font = Font(size=11, bold=False)

                    #Printing for debugging
                    print(id_value)
                    print(name_value)

                #Adding 1 to the index of where the program will write
                people_row_count += 1
            #Checking if there`s company and not mortgage in the line
            elif json_data['hebrew_Company'] in info and json_data['hebrew_Mortgage'] not in info:
                info += " "
                info = " " + info
                info = " ".join(info.split())
                print(info)

                if type_of_file == 1:
                    # Find the company ID and putting it in the excel
                    company_id_value = info[info.find(json_data['hebrew_Company']) - 10:info.find(json_data['hebrew_Company']) - 1]
                    sheet.cell(row=company_row_count, column=4).value = company_id_value
                    sheet.cell(row=company_row_count, column=4).font = Font(size=11, bold=False)

                    #Find the name and putting it in the excel by certain distance from the ID and the reason
                    company_name_value = info[info.find(json_data['hebrew_Company']) + 4:info.find(json_data['hebrew_Company']) + find_company_name_shared_homes(info)][::-1]
                    sheet.cell(row=company_row_count, column=5).value = company_name_value
                    sheet.cell(row=company_row_count, column=5).font = Font(size=11, bold=False)

                    #Printing for debugging
                    print(company_name_value)
                    print(company_id_value)

                if type_of_file == 2:
                    # Find the company ID and putting it in the excel
                    company_id_value = info[info.find(json_data['hebrew_Company']) - 10:info.find(json_data['hebrew_Company']) - 1]
                    sheet.cell(row=company_row_count, column=4).value = company_id_value
                    sheet.cell(row=company_row_count, column=4).font = Font(size=11, bold=False)

                    #Find the name and putting it in the excel by certain distance from the ID and the reason
                    company_name_value = info[info.find(json_data['hebrew_Company']) + 4:info.find(json_data['hebrew_Company']) + find_company_name_shared_rights(info)][::-1]
                    sheet.cell(row=company_row_count, column=5).value = company_name_value
                    sheet.cell(row=company_row_count, column=5).font = Font(size=11, bold=False)

                    #Printing for debugging
                    print(company_name_value)
                    print(company_id_value)

                #Adding 1 to the index of where the program will write
                company_row_count += 1
            #Checking if there`s passport and not mortgage in the line
            elif json_data['hebrew_passport'] in info:
                info += " "
                info = " " + info
                info = " ".join(info.split())
                print(info)

                if type_of_file == 1:
                    #Find the passport and putting it in the excel (more complicated check, ID comes in multiple lengths)
                    if " " in info[info.find(json_data['hebrew_passport']) - 9:info.find(json_data['hebrew_passport']) - 1]:
                        passport_value = info[info.find(json_data['hebrew_passport']) - 8:info.find(json_data['hebrew_passport']) - 1]
                    elif " " in info[info.find(json_data['hebrew_passport']) - 10:info.find(json_data['hebrew_passport']) - 1]:
                        passport_value = info[info.find(json_data['hebrew_passport']) - 9:info.find(json_data['hebrew_passport']) - 1]
                    elif " " in info[info.find(json_data['hebrew_passport']) - 11:info.find(json_data['hebrew_passport']) - 1]:
                        passport_value = info[info.find(json_data['hebrew_passport']) - 10:info.find(json_data['hebrew_passport']) - 1]
                    elif " " in info[info.find(json_data['hebrew_passport']) - 12:info.find(json_data['hebrew_passport']) - 1]:
                        passport_value = info[info.find(json_data['hebrew_passport']) - 11:info.find(json_data['hebrew_passport']) - 1]
                    elif " " in info[info.find(json_data['hebrew_passport']) - 13:info.find(json_data['hebrew_passport']) - 1]:
                        passport_value = info[info.find(json_data['hebrew_passport']) - 12:info.find(json_data['hebrew_passport']) - 1]
                    else:
                        passport_value = info[info.find(json_data['hebrew_passport']) - 13:info.find(json_data['hebrew_passport']) - 1]
                    sheet.cell(row=passport_row_count, column=7).value = passport_value
                    sheet.cell(row=passport_row_count, column=7).font = Font(size=11, bold=False)

                    # Find the name and putting it in the excel by certain distance from the passport and the reason
                    passport_name_value = info[info.find(json_data['hebrew_passport']) + 5:info.find(json_data['hebrew_passport']) + find_passport_name_shared_homes(info)][::-1]
                    sheet.cell(row=passport_row_count, column=8).value = passport_name_value
                    sheet.cell(row=passport_row_count, column=8).font = Font(size=11, bold=False)

                    #Printing for debugging
                    print(passport_value)
                    print(passport_name_value)
                #Adding 1 to the index of where the program will write
                passport_row_count += 1

    #Adding titles
    write_excel_titles(sheet)

    #Saving the excel
    book.save(excel_file)

    #Adding the information to the information file
    write_data_in_information_file(file_name)


def write_data_in_information_file(file_name):
    """writes in the information file the file name, and time of executing"""
    information_file = open("InformationFile.txt", "a")
    information_file.write("\n" + file_name + " " + str(date.today().strftime("%d/%m/%Y")) + " " + str(time.strftime("%H:%M:%S", time.localtime())))
    information_file.close()


def write_excel_titles(sheet):
    """gets a sheet and writes titles in it"""
    sheet.cell(row=1, column=1).value = "ת.ז"
    sheet.cell(row=1, column=1).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=2).value = "שם"
    sheet.cell(row=1, column=2).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=4).value = "מספר"
    sheet.cell(row=1, column=4).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=5).value = "שם חברה"
    sheet.cell(row=1, column=5).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=7).value = "מספר דרכון"
    sheet.cell(row=1, column=7).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=8).value = "שם"
    sheet.cell(row=1, column=8).font = Font(size=11, bold=True)


def clear_excel_cell(cell):
    """receive excel cell and clears it"""
    cell.value = None
    cell.font = Font(size=11, bold=False)
    cell.border = Border()


def write_file_type_in_excel(file_type, sheet):
    """gets sheet and file-type and writes it in the sheet"""
    sheet.cell(row=1, column=10).value = "סוג קובץ:"
    sheet.cell(row=1, column=10).font = Font(size=11, bold=True)
    sheet.cell(row=2, column=10).value = file_type
    sheet.cell(row=2, column=10).font = Font(size=11, bold=False)


def find_name_shared_rights(info):
    """Returning the distance of the name from the ID"""
    json_file = open('config.json', encoding="utf8")
    json_data = json.load(json_file)

    length = 3

    info = info[info.find(json_data['hebrew_ID']) + 3:]

    length += info.find(" ") + 1
    info = info[info.find(" ") + 1:]

    info = info[::-1]
    print(info)
    info = " ".join(info.split())
    info = info[name_reason_filtering_shared_rights(info):]

    info = " ".join(info.split())
    print(info)
    length += len(info)

    return length


def name_reason_filtering_shared_rights(info):
    """Returning the string without the reason in order to find the name"""
    json_file = open('config.json', encoding="utf8")
    json_data = json.load(json_file)

    index_of_reason = 0

    for reason in json_data['possible_name_reasons']:
        if reason in info:
            index_of_reason = info.find(reason) + len(reason)
            break
    json_file.close()
    return index_of_reason


def name_reason_filtering_shared_homes(info):
    """Returning the string without the reason in order to find the name"""

    json_file = open('config.json', encoding="utf8")
    json_data = json.load(json_file)

    for reason in json_data['possible_name_reasons']:
        if reason in info:
            info = info.replace(reason, "")
            break
    json_file.close()
    return info


def find_name_shared_homes(info):
    """Returning the distance of the name from the ID"""
    json_file = open('config.json', encoding="utf8")
    json_data = json.load(json_file)

    length = 3
    info = info[info.find(json_data['hebrew_ID']) + 3:]

    length += info.find(" ") + 1
    info = info[info.find(" ") + 1:]

    info = info[::-1]
    print(info)

    info = name_reason_filtering_shared_homes(info)

    info = " ".join(info.split())
    print(info)
    length += len(info)

    return length


def find_passport_name_shared_homes(info):
    """Returning the distance of the name from the passport ID"""
    json_file = open('config.json', encoding="utf8")
    json_data = json.load(json_file)

    length = 5
    info = info[info.find(json_data['hebrew_passport']) + 5:]

    info = " ".join(info.split())

    length += info.find(" ") + 1
    info = info[info.find(" ") + 1:]

    info = " ".join(info.split())

    info = info[::-1]
    print(info)

    info = name_reason_filtering_shared_homes(info)

    info = " ".join(info.split())
    print(info)
    length += len(info)

    return length


def find_company_name_shared_homes(info):
    """Returning the distance of the name from the company ID"""
    json_file = open('config.json', encoding="utf8")
    json_data = json.load(json_file)

    length = 4
    info = info[info.find(json_data['hebrew_Company']) + 4:]

    length += info.find(" ") + 1
    info = info[info.find(" ") + 1:]

    info = info[::-1]
    print(info)

    for reason in json_data['possible_company_name_reasons']:
        if reason in info:
            info = info.replace(reason, "")
    json_file.close()
    info = " ".join(info.split())

    print(info)
    length += len(info)

    return length


def find_company_name_shared_rights(info):
    """Returning the distance of the name from the company ID"""
    json_file = open('config.json', encoding="utf8")
    json_data = json.load(json_file)

    length = 4
    info = info[info.find(json_data['hebrew_Company']) + 4:]

    length += info.find(" ") + 1
    info = info[info.find(" ") + 1:]

    info = info[::-1]
    print(info)

    for reason in json_data['possible_company_name_reasons']:
        if reason in info:
            info = info[:info.find(reason)]
    json_file.close()
    info = " ".join(info.split())

    print(info)
    length += len(info)

    return length


def find_file_type(info, sheet):
    """Returning the type of the file presented by numbers"""
    if "םיפתושמ םיתב" in info:
        print(info)
        write_file_type_in_excel("בתים משותפים", sheet)
        return 1
    elif "תויוכזה סקנפמ" in info:
        print(info)
        write_file_type_in_excel("פנקס זכויות", sheet)
        return 2
    else:
        return 0


pdf_to_txt('143.pdf')
#print(find_passport_name_shared_homes(info))
#print(info[info.find("ןוכרד") - 10:info.find("ןוכרד") - 1])
#print(info[info.find("ןוכרד") + 5:info.find("ןוכרד") + find_passport_name_shared_homes(info)][::-1])


